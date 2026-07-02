"""
Gestion du fichier Excel historique du COSEC (historique_cosec.xlsx).

Depuis le 23/06/2026, ce classeur heberge TROIS historiques mensuels
independants, chacun dans son propre onglet :
  - "Typologies"    : une ligne par (Mois, Typologie) -- cf write_history()
  - "Surveillance"  : une ligne par Mois (gravite, cloture, MTTA/MTTR/MTTC)
                       -- cf write_surveillance_history()
  - "SLA"           : une ligne par incident en depassement de SLA
                       (MTTA et/ou MTTR) pour le mois -- cf
                       write_sla_history()

(L'onglet "Typologies" s'appelait "Historique" avant le renommage du
22/06/2026 ; le nom a ete aligne sur celui de l'onglet "Surveillance" pour
la clarte, le classeur portant desormais plusieurs historiques.)

Idempotent par mois dans les trois cas : relancer l'ecriture pour un mois
deja present REMPLACE ses lignes (pas de doublon), les autres mois sont
preserves -- y compris quand le nouveau resultat est une liste VIDE (ex:
aucun depassement SLA ce mois-ci doit pouvoir effacer un ancien resultat
errone pour ce meme mois). Chaque fonction de lecture/ecriture ne touche
QUE son propre onglet -- l'ordre d'appel entre write_history(),
write_surveillance_history() et write_sla_history() n'a donc aucune
importance.
"""

import json
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

SHEET_NAME = "Typologies"
HEADERS = ["Mois", "Typologie", "Sources d'alerte", "Nombre d'incidents"]
TABLE_NAME = "HistoriqueTypologies"


def _new_workbook():
    """Cree un classeur Excel vide avec l'onglet Typologies et sa ligne d'en-tete."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    return wb, ws


def _read_existing_rows(ws) -> list:
    """Relit les lignes de donnees (hors en-tete) d'une feuille Typologies
    deja existante, sous forme de dicts {Month, Title, AlertSources,
    IncidentCount} -- utilise avant remplacement pour preserver les mois
    non concernes par l'ecriture en cours."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append({"Month": row[0], "Title": row[1], "AlertSources": row[2], "IncidentCount": row[3]})
    return rows


def _to_display_sources(sources) -> str:
    """Convertit une chaine JSON (['a','b']) en texte 'a, b' ; laisse tel quel sinon."""
    if isinstance(sources, str) and sources.strip().startswith("["):
        try:
            return ", ".join(json.loads(sources))
        except (TypeError, ValueError):
            return sources
    return sources or ""


def _apply_formatting(ws, n_rows: int):
    """Applique la mise en forme de l'onglet Typologies : en-tete blanc sur
    fond bleu marque, corps en Arial 10, largeurs de colonnes fixes et
    figeage de la ligne d'en-tete (freeze_panes)."""
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    body_font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = body_font
        row[1].alignment = Alignment(wrap_text=False)

    for i, w in enumerate([10, 60, 38, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def write_history(path: str, new_rows: list):
    """
    Ecrit/met a jour l'onglet Historique avec les lignes agregees fournies
    (sortie de typology_normalize.aggregate_typology_rows).

    new_rows : liste de dicts {Month, Title, AlertSources, IncidentCount}
    """
    months_to_replace = {r["Month"] for r in new_rows}
    existing = []

    if os.path.exists(path):
        wb = load_workbook(path)
        if SHEET_NAME in wb.sheetnames:
            existing = _read_existing_rows(wb[SHEET_NAME])
            existing = [r for r in existing if r["Month"] not in months_to_replace]
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME, 0)
        ws.append(HEADERS)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 \
                and wb["Sheet"]["A1"].value is None:
            del wb["Sheet"]
    else:
        wb, ws = _new_workbook()

    all_rows = existing + new_rows
    all_rows.sort(key=lambda r: (r["Month"], -int(r["IncidentCount"])))

    for r in all_rows:
        ws.append([r["Month"], r["Title"], _to_display_sources(r["AlertSources"]), int(r["IncidentCount"])])

    n_rows = len(all_rows)
    _apply_formatting(ws, n_rows)

    table_ref = f"A1:D{n_rows + 1}"
    table = Table(displayName=TABLE_NAME, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    ws.add_table(table)

    wb.save(path)
    return n_rows


# ---------------------------------------------------------------------------
# Onglet "Surveillance" (gravite / cloture / MTTA-MTTR-MTTC) -- une ligne
# par MOIS (et non par typologie : il n'y a qu'un nombre fixe de mesures
# par mois, pas de liste a developper -- cf surveillance_normalize.py).
# ---------------------------------------------------------------------------

SURVEILLANCE_SHEET_NAME = "Surveillance"

# Ordre des colonnes = ordre des cles attendues dans le dict passe a
# write_surveillance_history() (cf surveillance_normalize.build_surveillance_row).
SURVEILLANCE_COLUMNS = [
    ("Month", "Mois"),
    ("Total", "Total incidents"),
    ("High", "Gravité - Élevée"),
    ("Medium", "Gravité - Moyenne"),
    ("Low", "Gravité - Faible"),
    ("Informational", "Gravité - Informationnelle"),
    ("TruePositive", "Clôture - Vrai positif"),
    ("FalsePositive", "Clôture - Faux positif"),
    ("BenignPositive", "Clôture - Positif bénin"),
    ("Undetermined", "Clôture - Indéterminé"),
    ("MTTA", "MTTA (h)"),
    ("MTTR", "MTTR (h)"),
    ("MTTC", "MTTC (h)"),
]
SURVEILLANCE_HEADERS = [label for _, label in SURVEILLANCE_COLUMNS]
SURVEILLANCE_KEYS = [key for key, _ in SURVEILLANCE_COLUMNS]
SURVEILLANCE_TABLE_NAME = "HistoriqueSurveillance"

# Colonnes entieres vs decimales (MTTA/MTTR/MTTC sont arrondies a 3 decimales,
# le reste sont des comptes d'incidents).
_SURVEILLANCE_FLOAT_KEYS = {"MTTA", "MTTR", "MTTC"}


def _read_existing_surveillance_rows(ws) -> list:
    """Relit les lignes de donnees (hors en-tete) d'une feuille Surveillance
    deja existante, sous forme de dicts cles sur SURVEILLANCE_KEYS -- une
    ligne par mois deja present dans l'historique."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(SURVEILLANCE_KEYS, row)))
    return rows


def _apply_surveillance_formatting(ws, n_rows: int):
    """Applique la mise en forme de l'onglet Surveillance : memes couleurs
    de marque que l'onglet Typologies, mais toutes les colonnes (y compris
    les mesures) sont centrees -- il n'y a pas de colonne "libelle long"
    a aligner a gauche comme la Typologie."""
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    body_font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center")

    widths = [10] + [16] * (len(SURVEILLANCE_HEADERS) - 1)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def write_surveillance_history(path: str, month_row: dict) -> int:
    """
    Ecrit/met a jour l'onglet Surveillance avec la ligne agregee fournie
    pour UN mois (sortie de surveillance_normalize.build_surveillance_row).
    Idempotent par mois : remplace la ligne existante pour ce mois si elle
    existe, sinon l'ajoute ; les autres mois sont preserves.

    month_row : dict avec les cles Month, Total, High, Medium, Low,
                Informational, TruePositive, FalsePositive, BenignPositive,
                Undetermined, MTTA, MTTR, MTTC.

    Ne touche QUE l'onglet "Surveillance" -- n'affecte pas l'onglet
    "Typologies" s'il existe deja dans le meme classeur.
    """
    month = month_row["Month"]
    existing = []

    if os.path.exists(path):
        wb = load_workbook(path)
        if SURVEILLANCE_SHEET_NAME in wb.sheetnames:
            existing = _read_existing_surveillance_rows(wb[SURVEILLANCE_SHEET_NAME])
            existing = [r for r in existing if r["Month"] != month]
            del wb[SURVEILLANCE_SHEET_NAME]
        ws = wb.create_sheet(SURVEILLANCE_SHEET_NAME)
        ws.append(SURVEILLANCE_HEADERS)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 \
                and wb["Sheet"]["A1"].value is None:
            del wb["Sheet"]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SURVEILLANCE_SHEET_NAME
        ws.append(SURVEILLANCE_HEADERS)

    all_rows = existing + [month_row]
    all_rows.sort(key=lambda r: r["Month"])

    for r in all_rows:
        line = []
        for key in SURVEILLANCE_KEYS:
            if key == "Month":
                line.append(r[key])
            elif key in _SURVEILLANCE_FLOAT_KEYS:
                value = r.get(key)
                line.append(round(float(value), 3) if value is not None else None)
            else:
                line.append(int(r.get(key) or 0))
        ws.append(line)

    n_rows = len(all_rows)
    _apply_surveillance_formatting(ws, n_rows)


    wb.save(path)
    return n_rows


# ---------------------------------------------------------------------------
# Onglet "SLA" (depassements MTTA / MTTR) -- une ligne par INCIDENT en
# depassement pour le mois (et non une ligne par mois comme l'onglet
# Surveillance : un mois peut avoir 0, 1 ou N depassements -- cf
# sla_normalize.build_sla_rows).
# ---------------------------------------------------------------------------

SLA_SHEET_NAME = "SLA"
SLA_HEADERS = ["Mois", "Type SLA", "N°INC", "Sévérité", "Titre", "Créé le", "Attribution", "Clôture"]
SLA_KEYS = ["Month", "TypeSLA", "IncidentNumber", "Severity", "Title",
            "CreatedTime", "AttributionTime", "ClosedTime"]
SLA_TABLE_NAME = "HistoriqueSLA"

# Colonnes date (index 1-based dans la ligne Excel) -- Créé le / Attribution
# / Clôture. Clôture peut etre vide (incident MTTA pas encore clôturé).
_SLA_DATE_COLUMNS = (6, 7, 8)
_SLA_DATE_FORMAT = "DD/MM/YYYY HH:MM"


def _read_existing_sla_rows(ws) -> list:
    """Relit les lignes de donnees (hors en-tete) d'une feuille SLA deja
    existante, sous forme de dicts cles sur SLA_KEYS -- une ligne par
    incident en depassement, tous mois confondus."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(SLA_KEYS, row)))
    return rows


def _apply_sla_formatting(ws, n_rows: int):
    """Applique la mise en forme de l'onglet SLA : memes couleurs de marque
    que les deux autres onglets, format de date DD/MM/YYYY HH:MM sur les
    3 colonnes de dates (cf _SLA_DATE_COLUMNS), et centrage des colonnes
    courtes (Type SLA, N°INC, Sévérité)."""
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    body_font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = body_font
        for col_idx in _SLA_DATE_COLUMNS:
            cell = row[col_idx - 1]
            if cell.value is not None:
                cell.number_format = _SLA_DATE_FORMAT
        row[2].alignment = Alignment(horizontal="center")  # N°INC
        row[1].alignment = Alignment(horizontal="center")  # Type SLA
        row[3].alignment = Alignment(horizontal="center")  # Sévérité

    widths = [10, 11, 11, 13, 55, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def write_sla_history(path: str, year: int, month: int, new_rows: list) -> int:
    """
    Ecrit/met a jour l'onglet SLA avec les incidents en depassement (MTTA
    et/ou MTTR) pour le mois CIBLE donne (sortie de
    sla_normalize.build_sla_rows).

    Idempotent par mois : remplace TOUTES les lignes existantes du mois
    cible -- meme si new_rows est une liste VIDE (aucun depassement ce
    mois-ci doit pouvoir effacer un ancien resultat errone pour ce meme
    mois) --, les autres mois sont preserves.

    new_rows : liste de dicts {Month, TypeSLA, IncidentNumber, Severity,
               Title, CreatedTime, AttributionTime, ClosedTime} (dates en
               objets datetime ou None).

    Retourne le nombre de depassements ecrits pour ce mois (len(new_rows),
    PAS le total toutes lignes confondues -- contrairement a write_history/
    write_surveillance_history, car "nombre de depassements ce mois" est
    l'information utile a logger ici, y compris 0).
    """
    month_str = f"{year:04d}-{month:02d}"
    existing = []

    if os.path.exists(path):
        wb = load_workbook(path)
        if SLA_SHEET_NAME in wb.sheetnames:
            existing = _read_existing_sla_rows(wb[SLA_SHEET_NAME])
            existing = [r for r in existing if r["Month"] != month_str]
            del wb[SLA_SHEET_NAME]
        ws = wb.create_sheet(SLA_SHEET_NAME)
        ws.append(SLA_HEADERS)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 \
                and wb["Sheet"]["A1"].value is None:
            del wb["Sheet"]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SLA_SHEET_NAME
        ws.append(SLA_HEADERS)

    all_rows = existing + new_rows
    all_rows.sort(key=lambda r: (r["Month"], r.get("CreatedTime") or datetime.min))

    for r in all_rows:
        ws.append([r.get(key) for key in SLA_KEYS])

    n_rows = len(all_rows)
    _apply_sla_formatting(ws, n_rows)

    # Pas de Table Excel native si la feuille n'a aucune ligne de donnees
    # (classeur tout juste cree, aucun mois encore traite) -- une Table sur
    # une plage reduite a l'en-tete seul ("A1:H1") n'est pas valide. Un mois
    # traite SANS depassement, lui, laisse n_rows a 0 UNIQUEMENT si c'est le
    # tout premier mois jamais traite ; dans tous les autres cas, les lignes
    # des AUTRES mois (existing) maintiennent n_rows > 0.
    if n_rows > 0:
        last_col = get_column_letter(len(SLA_HEADERS))
        table_ref = f"A1:{last_col}{n_rows + 1}"
        table = Table(displayName=SLA_TABLE_NAME, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
        ws.add_table(table)

    wb.save(path)
    return len(new_rows)
